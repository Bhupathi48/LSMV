import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

def filter_and_copy_excel(excel_path, source_sheet, target_sheet, filter_column, filter_value):
    # Read the source sheet
    df = pd.read_excel(excel_path, sheet_name=source_sheet)
    # Filter rows
    filtered_df = df[df[filter_column] == filter_value]
    # Write filtered data to new sheet
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        filtered_df.to_excel(writer, sheet_name=target_sheet, index=False)
    # Autofit columns and color header row
    wb = load_workbook(excel_path)
    ws = wb[target_sheet]
    # Autofit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    # Color header row (first row) in light blue
    header_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    for cell in ws[1]:
        cell.fill = header_fill
    # Add thin borders to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    wb.save(excel_path)
    print(f"Filtered data written to sheet '{target_sheet}' in '{excel_path}' with autofit columns, colored header, and borders.")

if __name__ == "__main__":
    # excel_path = os.path.join(os.path.dirname(__file__), '../Tests/Web/POC/LSMV_TestData.xlsx')
    # excel_path = os.path.abspath(excel_path)
    # source_sheet = 'CreateRCT1'
    # target_sheet = 'CreateRCT_Exe'
    # filter_column = 'Execution'
    # filter_value = "Yes"
    # filter_and_copy_excel(excel_path, source_sheet, target_sheet, filter_column, filter_value)
