"""
ExcelTestSelector.py - Excel-based test execution control for LSMV CreateRCT1 sheet

This library controls test case execution based on the CreateRCT1 sheet in LSMV_TestData.xlsx.
It reads from Column 2 (Scenario Name) and Column 8 (Execution) to determine which tests to run.
"""

import os
import openpyxl
from robot.api import logger


class ExcelTestSelector:
    """
    Excel-based test selector for LSMV CreateRCT1 sheet
    Reads Scenario Name (Col 2) and Execution flag (Col 8) from CreateRCT1 sheet
    """
    
    def __init__(self):
        self.current_directory = os.path.dirname(os.path.realpath(__file__))
        self.parent_directory = os.path.abspath(os.path.join(self.current_directory, os.pardir))
        self._executable_tests_cache = None  # Cache for loaded Excel config
        self._config_logged = False  # Flag to track if configuration has been logged
        
    def _load_excel_config(self, excel_file_name, sheet_name):
        """Load and parse Excel configuration from CreateRCT1 sheet structure"""
        # For CreateRCT1 sheet, look in Resource/TestData_Excels directory
        excel_path = os.path.join(self.parent_directory, "Resource", "TestData_Excels", excel_file_name)
        
        if not os.path.exists(excel_path):
            logger.warn(f"Excel file not found: {excel_path}. All tests will execute.")
            return []
        
        try:
            workbook = openpyxl.load_workbook(excel_path)
            if sheet_name not in workbook.sheetnames:
                logger.warn(f"Sheet '{sheet_name}' not found. All tests will execute.")
                return []
                
            worksheet = workbook[sheet_name]
            executable_tests = []
            
            # CreateRCT1 structure: Col 2 = Scenario Name, Col 8 = Execution
            for row in range(2, worksheet.max_row + 1):
                scenario_name = worksheet.cell(row, 2).value  # Column 2: Scenario Name
                execute_flag = worksheet.cell(row, 8).value   # Column 8: Execution
                test_case_id = worksheet.cell(row, 1).value   # Column 1: Test Case ID (for logging)
                
                if scenario_name and str(execute_flag).lower() in ['yes', 'y', 'true', '1']:
                    executable_tests.append(scenario_name.strip())
                    logger.debug(f"Test enabled: {test_case_id} - {scenario_name}")
                elif scenario_name:
                    logger.debug(f"Test disabled: {test_case_id} - {scenario_name}")
            
            logger.info(f"CreateRCT1 configuration loaded: {len(executable_tests)} tests enabled out of {worksheet.max_row - 1} total")
            return executable_tests
            
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            return []
    
    def check_test_execution_allowed(self, excel_file_name="LSMV_TestData.xlsx", sheet_name="CreateRCT1"):
        """
        Check if current test should be executed based on CreateRCT1 Excel configuration
        This loads Excel config on first call and caches it for subsequent tests
        """
        from robot.libraries.BuiltIn import BuiltIn
        
        builtin = BuiltIn()
        
        # Load Excel config if not already cached
        if self._executable_tests_cache is None:
            self._executable_tests_cache = self._load_excel_config(excel_file_name, sheet_name)
            # Log configuration only once per process - concise version
            if not self._config_logged:
                logger.info(f"LSMV Excel Test Selector: {len(self._executable_tests_cache)} tests enabled from {sheet_name} sheet")
                self._config_logged = True
        
        executable_tests = self._executable_tests_cache
        
        # Get the current test name
        test_name = builtin.get_variable_value("${TEST_NAME}")
        
        if not executable_tests:
            logger.warn("No CreateRCT1 configuration found. Test will execute.")
            return
        
        # Try to match test name with scenario names from CreateRCT1
        matched_scenario = self._find_matching_scenario(test_name, executable_tests)
        
        if matched_scenario:
            logger.info(f"Test '{test_name}' matches scenario '{matched_scenario}' - Proceeding.")
        else:
            logger.info(f"Test '{test_name}' not found in CreateRCT1 execution list. Skipping.")
            builtin.skip(f"Test '{test_name}' not enabled in CreateRCT1 sheet")
    
    def _find_matching_scenario(self, test_name, executable_scenarios):
        """
        Find matching scenario name for Robot Framework test name
        Handles various naming conventions and partial matches
        """
        # Direct match first
        if test_name in executable_scenarios:
            return test_name
        
        # Try to find partial matches for common LSMV test patterns
        for scenario in executable_scenarios:
            # Check if scenario name is contained in test name or vice versa
            if scenario.lower() in test_name.lower() or test_name.lower() in scenario.lower():
                return scenario
            
            # Handle PULSE-CDDR pattern matching
            if "pulse-cddr" in test_name.lower() and "pulse-cddr" in scenario.lower():
                # Extract CDDR number from both
                test_cddr = self._extract_cddr_number(test_name)
                scenario_cddr = self._extract_cddr_number(scenario)
                if test_cddr and scenario_cddr and test_cddr == scenario_cddr:
                    return scenario
            
            # Handle special cases like DueDateRule
            if "duedaterule" in test_name.lower() and "duedaterule" in scenario.lower():
                return scenario
        
        return None
    
    def _extract_cddr_number(self, text):
        """
        Extract CDDR number from test/scenario name for matching
        E.g., 'LSMV_Autocalc_PULSE_CDDR91' -> '91'
        """
        import re
        match = re.search(r'cddr[-_]?(\d+)', text.lower())
        return match.group(1) if match else None


# Global instance
excel_test_selector = ExcelTestSelector()


def check_test_execution_allowed(excel_file_name="LSMV_TestData.xlsx", sheet_name="CreateRCT1"):
    """
    Robot Framework keyword: Check if current test should execute based on CreateRCT1 sheet
    Call this in Test Setup to load Excel config and skip tests not marked for execution
    
    Args:
        excel_file_name: Excel file name (default: LSMV_TestData.xlsx)
        sheet_name: Sheet name (default: CreateRCT1)
    """
    return excel_test_selector.check_test_execution_allowed(excel_file_name, sheet_name)