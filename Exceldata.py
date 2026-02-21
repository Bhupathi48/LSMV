import openpyxl
import os
from openpyxl import load_workbook
import pandas as pd
def path(file):
    fileDir = os.path.dirname(os.path.realpath('__file__'))
    filepath = os.path.join(fileDir, file)
    return filepath

# // Write a function to get entire data from the specific sheet
def get_entire_data_from_excel(filename, sheetname):
    workbook = load_workbook(filename)
    worksheet = workbook.get_sheet_by_name(sheetname)
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(list(row))
    return data

def get_entire_data_as_dict_from_excel(filename, sheetname):
    workbook = load_workbook(filename)
    worksheet = workbook.get_sheet_by_name(sheetname)
    data = []
    headers = [cell.value for cell in worksheet[1]]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers))}
        data.append(row_data)
    return data

def read_scenario_from_sheet(filename, sheetname, filter_column, filter_value):
    # Load the workbook and select the sheet
    workbook = load_workbook(filename)
    worksheet = workbook[sheetname]
    
    # Get headers
    headers = [cell.value for cell in worksheet[1]]
    
    # Find index of the filter_column
    if filter_column not in headers:
        raise ValueError(f"Column '{filter_column}' not found in sheet '{sheetname}'")
    col_idx = headers.index(filter_column)
    
    # Iterate through rows to find the match
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[col_idx] == filter_value:
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            return row_dict    
 # If no match found
    return None

def write_value_to_scenario(file_path, sheet_name, scenario_column, scenario_value, target_column, new_value):
    xls = pd.ExcelFile(file_path)

    sheets = {sheet: xls.parse(sheet) for sheet in xls.sheet_names}
    df = sheets[sheet_name]
    if scenario_column not in df.columns:
        raise ValueError(f"Column '{scenario_column}' not found in sheet '{sheet_name}'")
    if target_column not in df.columns:
        raise ValueError(f"Target column '{target_column}' not found in sheet '{sheet_name}'")
    match_idx = df[df[scenario_column] == scenario_value].index
    if not match_idx.empty:
        df.at[match_idx[0], target_column] = new_value
        # Save back all sheets (to preserve other sheets)
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            for name, data in sheets.items():
                if name == sheet_name:
                    data = df
                data.to_excel(writer, sheet_name=name, index=False)
        return True
    else:
        return False

# excel_path = "C:\\Users\\HPotdar\\JnJ_Project\\Latest_code\\lsmv_automation_poc\\LSMV\\Tests\\Web\\POC\\LSMV_TestData.xlsx"

# write_value_to_scenario(excel_path, "CreateRCT1", "Scenario Name", "PULSE-0010-Scenario4", "RCT", "RCT20250026077")

def update_cell_by_excel_row(file_path, sheet_name, excel_row_number, column_name, new_value):
    """
    Updates a cell in the given sheet by Excel row number (as seen in Excel, starting from 1).
    Highlights the cell green for 'pass' and red for 'fail'.
    :param file_path: Path to the Excel file
    :param sheet_name: Name of the sheet to update
    :param excel_row_number: The row number as seen in Excel (header is row 1, first data row is 2)
    :param column_name: The column name to update
    :param new_value: The value to write ('pass' or 'fail')
    :return: True if update is successful, False otherwise
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    # Convert Excel row number to pandas DataFrame index
    excel_row_number = int(excel_row_number)
    excel_row_number += 1
    row_number = excel_row_number - 2  # Excel row 2 is index 0, Excel row 3 is index 1, etc.
    xls = pd.ExcelFile(file_path)
    sheets = {sheet: xls.parse(sheet, keep_default_na=True, na_values=['']) for sheet in xls.sheet_names}
    df = sheets[sheet_name]
    # Create column at the end if it doesn't exist
    if column_name not in df.columns:
        # Add the new column at the end
        df[column_name] = ''
    if row_number < 0 or row_number >= len(df):
        raise IndexError(f"Excel row number {excel_row_number} is out of range for sheet '{sheet_name}'")
    df.at[row_number, column_name] = new_value
    # Save back all sheets (to preserve other sheets)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        for name, data in sheets.items():
            if name == sheet_name:
                data = df
            data.to_excel(writer, sheet_name=name, index=False, na_rep='NA')

    # Highlight the cell with color
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    col_idx = list(df.columns).index(column_name) + 1  # openpyxl is 1-based
    cell = ws.cell(row=excel_row_number, column=col_idx)
    if str(new_value).lower() == "pass":
        cell.fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")  # Light green
    elif str(new_value).lower() == "fail":
        cell.fill = PatternFill(start_color="ff6347", end_color="ff6347", fill_type="solid")  # Red
    wb.save(file_path)
    return True

def update_rct_by_excel_row(file_path, sheet_name, excel_row_number, rct_column_name, rct_value):
    """
    Updates the RCT number in the given sheet by Excel row number (as seen in Excel, starting from 1).
    :param file_path: Path to the Excel file
    :param sheet_name: Name of the sheet to update
    :param excel_row_number: The row number as seen in Excel (header is row 1, first data row is 2)
    :param rct_column_name: The column name containing RCT numbers (e.g., 'RCT')
    :param rct_value: The RCT number to write
    :return: True if update is successful, False otherwise
    """
    from openpyxl import load_workbook
    
    # Convert Excel row number to pandas DataFrame index
    excel_row_number = int(excel_row_number)
    row_number = excel_row_number - 2  # Excel row 2 is index 0, Excel row 3 is index 1, etc.
    
    xls = pd.ExcelFile(file_path)
    sheets = {sheet: xls.parse(sheet) for sheet in xls.sheet_names}
    df = sheets[sheet_name]
    # Create column at the start if it doesn't exist
    if rct_column_name not in df.columns:
        # Insert the new column at the beginning (index 0)
        df.insert(0, rct_column_name, '')
    if row_number < 0 or row_number >= len(df):
        raise IndexError(f"Excel row number {excel_row_number} is out of range for sheet '{sheet_name}'")
    
    # Update the RCT value
    df.at[row_number, rct_column_name] = rct_value
    
    # Save back all sheets (to preserve other sheets)
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        for name, data in sheets.items():
            if name == sheet_name:
                data = df
            data.to_excel(writer, sheet_name=name, index=False)
    
    return True

def copy_sheet_to_new_excel(file_path, sheet_name, status_column=None, status_value=None, output_folder=None, output_filename=None):
    """
    Copies a specific sheet from the Excel file and creates a new Excel file with that sheet.
    The new file will be named as <output_filename> if provided, else <sheet_name>.xlsx. Optionally, mark a status in a given column for all rows.
    Adds a new column 'Actual Test Result' at the end of the sheet.
    :param file_path: Path to the source Excel file
    :param sheet_name: Name of the sheet to copy
    :param status_column: (Optional) Column name to mark status
    :param status_value: (Optional) Value to set in the status column for all rows
    :param output_folder: (Optional) Folder path to save the new file. If None, uses the source file's folder.
    :param output_filename: (Optional) Custom name for the output Excel file. If None, uses <sheet_name>.xlsx
    :return: Path to the new Excel file
    """
    import os
    from filelock import FileLock, Timeout
    lock_path = file_path + ".lock"
    with FileLock(lock_path, timeout=60):
        xls = pd.ExcelFile(file_path)
        if sheet_name not in xls.sheet_names:
            raise ValueError(f"Sheet '{sheet_name}' not found in file '{file_path}'")
        df = xls.parse(sheet_name, keep_default_na=True, na_values=[''])
        if status_column and status_value is not None:
            if status_column in df.columns:
                df[status_column] = status_value
            else:
                raise ValueError(f"Status column '{status_column}' not found in sheet '{sheet_name}'")
        # Add new column 'Actual Test Result' at the end
        df['Actual Test Result'] = ''
        # Determine output folder
        if output_folder is None:
            output_folder = os.path.dirname(file_path)
        os.makedirs(output_folder, exist_ok=True)
        # Determine output filename
        if output_filename is None:
            new_file_name = f"{sheet_name}.xlsx"
        else:
            new_file_name = output_filename if output_filename.endswith('.xlsx') else output_filename + '.xlsx'
        new_file_path = os.path.join(output_folder, new_file_name)
        # Delete if file already exists
        if os.path.exists(new_file_path):
            os.remove(new_file_path)
        # Save DataFrame to Excel while preserving NaN values
        with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, na_rep='NA')

        # Adjust column widths using openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(new_file_path)
        ws = wb[sheet_name]
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Add some padding
            ws.column_dimensions[col_letter].width = adjusted_width
        wb.save(new_file_path)
        return new_file_path
# excel_path = "C:\\Users\\HPotdar\\JnJ_Project\\Latest_code\\lsmv_automation_poc\\LSMV\\Tests\\Web\\POC\\LSMV_TestData.xlsx"
# copy_sheet_to_new_excel(excel_path, "PULSE-0027-Scenario1", output_folder="C:\\Users\\HPotdar\\JnJ_Project\\Latest_code\\lsmv_automation_poc\\LSMV\\Output")

def get_actual_test_results(excel_path):
    df = pd.read_excel(excel_path)
    # Display the first few rows
    print(df.head())
    # Fetch the 'Actual Test Result' column as a list
    if 'Actual Test Result' in df.columns:
        actual_test_results = df['Actual Test Result'].tolist()
        print(actual_test_results)
        failed_rows = df[df['Actual Test Result'] == 'FAIL']
        if not failed_rows.empty:
            print("Rows where test failed:")
            print(failed_rows)
            # Optionally, return the failed rows as a list of dicts
            return failed_rows.to_dict(orient='records'), False
        else:
            print("No failures found.")
            return f"No failures found.",True
    else:
        print("Column 'Actual Test Result' not found in the Excel file.")
        return None